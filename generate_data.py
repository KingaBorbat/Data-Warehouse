# Generates and prepares data for database insertion


import random
import datetime
import string
import mysql.connector
from mysql.connector import Error


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

categories = ["top", "bottom", "footwear", "dresses", "beachwear", "activewear"]
subcategories = [
    ["top", "shirt", "t-shirts", "blouse", "hanorac", "jacket"],
    ["pant", "jeans", "sweatpant", "shorts", "skirt"],
    ["sneakers", "tenis", "sandal", "slippers", "casual footwear"],
    [
        "mini dress",
        "midi dress",
        "long dress",
        "casual dress",
        "boho  dress",
        "elegant  dress",
    ],
    ["beachwear top", "beachwear bottom", "beachwear set", "beach cover-up"],
    [
        "sport jacket",
        "sport sweatshirt",
        "sport t-shirt",
        "sport pants",
        "sport leggings",
        "sport shorts",
    ],
]

genders = ["man", "woman"]
sizes = ["XS", "S", "M", "L", "XL", "XXL", "XXXL", "XXXXL"]
fabrics = ["polyester", "cotton", "silk", "leather", "mesh", "stain", "denim"]
countries = ["Romania"]
counties = ["Cluj", "Maros", "Harghita", "Covasna"]
locations = [
    ["Cluj-Napoca", "Floresti"],
    ["Targu-Mures", "Sighisoara"],
    [
        "Miercurea Ciuc",
        "Odorheiu Secuiesc",
    ],
    [
        "Sfantu Gheorghe",
        "Targu Secuiesc",
    ],
]

colors = [
    "red",
    "yellow",
    "blue",
    "green",
    "brown",
    "orange",
    "black",
    "white",
    "purple",
]
patterns = [
    "plain",
    "letters",
    "floral",
    "striped",
    "graphic",
    "checks",
    "bohemian",
    "botanical",
]

adjectives = [
    "Comfy",
    "Stylish",
    "Classic",
    "Chic",
    "Trendy",
    "Cozy",
    "Fashionable",
]


counties_to_insert = []
fabrics_to_insert = []
customer_locations = []
categories_to_insert = []
subcategories_to_insert = []
locations_to_insert = []
customer_locations_to_insert = []
stores_to_insert = []
customers_to_insert = []
products_to_insert = []
fabric_details_to_insert = []
sales_to_insert = []
sale_details_to_insert = []
customer_ids = []
product_ids = []
location_ids = []

# Generates a specified number of unique numeric IDs of a given length.
def generate_id(num_id, length_id, list=[]):
    ids = []
    while len(ids) < num_id:
        id = str(random.randint(1, 9))
        for _ in range(length_id - 1):
            id += str(random.randint(0, 9))
        id = int(id)
        if id not in ids and id not in list:
            ids.append(id)
    return ids


# Prepares the list of categories for database insertion
def generate_categories(categories):
    nr_cat = 1
    categories_ = []
    for i in range(len(categories)):
        if categories[i] == "dresses":
            categories_.append((nr_cat, categories[i], genders[1]))
            nr_dresses = nr_cat
            nr_cat += 1

        else:
            categories_.append((nr_cat, categories[i], genders[0]))
            nr_cat += 1
            categories_.append((nr_cat, categories[i], genders[1]))
            nr_cat += 1
    return nr_dresses, categories_


# Prepares the list of subcategories for database insertion
def generate_subcategories(subcategories, nr_dresses, categories_):
    subcategories_ = []
    nr_subcat = 1
    n = 1
    m = 0

    while n < len(categories_):
        for s in subcategories[m]:
            if n == nr_dresses:
                subcategories_.append((nr_subcat, s, n))
                nr_subcat += 1
            else:
                subcategories_.append((nr_subcat, s, n))
                nr_subcat += 1
                subcategories_.append((nr_subcat, s, n + 1))
                nr_subcat += 1
        if n == nr_dresses:
            n += 1
        else:
            n += 2
        m += 1
    return subcategories_


# Prepares the list of store and customer locations for database insertion
def generate_locations(counties, locations, customer_locations, store_location_ids, customer_location_ids):
    locations_ = []
    customer_locations_ = []
    loc_id = 0

    for i in range(1, len(counties) + 1):
        for l in locations[i - 1]:
            locations_.append((store_location_ids[loc_id], l, i))
            customer_locations_.append((store_location_ids[loc_id], l, i))
            loc_id += 1

    loc_id = 0
    for i in range(len(customer_locations)):
        rand_county = random.randint(1, len(counties))
        customer_locations_.append(
            (customer_location_ids[loc_id], customer_locations[i], rand_county)
        )
        loc_id += 1
    return locations_, customer_locations_


# Generates stores and prepares data for database insertion
def generate_stores(locations, store_location_ids):
    nr_store = 1
    nr_stores_in_city = [2, 1]
    loc_id = 0
    stores_ = []
    for i in range(0, len(locations)):
        for j in range(0, len(locations[i])):
            for k in range(nr_stores_in_city[j]):
                stores_.append(
                    (
                        nr_store,
                        "store" + str(nr_store),
                        store_location_ids[loc_id],
                        "address" + str(nr_store),
                    )
                )
                nr_store += 1
            loc_id += 1
    return stores_


# Generates an email
def generate_email():
    letters = string.ascii_lowercase
    return "".join(random.choice(letters) for _ in range(8)) + "@example.com"


# Generates a phone number
def generate_phone_number():
    return "".join(random.choices(string.digits, k=10))


# Generates customers and prepares data for database insertion
def generate_customers(nr_customers, customer_ids, location_ids, i=0):
    customers = []

    random.shuffle(location_ids)
    n = 0
    if nr_customers > len(location_ids):
        n = len(location_ids)
    else:
        n = nr_customers
    while i < n:
        customer_name = f"customer{i+1}"
        email = generate_email()
        phone_number = generate_phone_number()
        gender = random.choice(genders)
        size = random.choice(sizes)
        print(i)
        customers.append(
            (
                customer_ids[i],
                customer_name,
                email,
                phone_number,
                gender,
                location_ids[i],
                size,
            )
        )
        i += 1

    if nr_customers > len(location_ids):
        while i < nr_customers:
            customer_name = f"customer{i}"
            email = generate_email()
            phone_number = generate_phone_number()
            gender = random.choice(genders)
            size = random.choice(sizes)
            location_id = random.choice(location_ids)
            customers.append(
                (
                    customer_ids[i],
                    customer_name,
                    email,
                    phone_number,
                    gender,
                    location_id,
                    size,
                )
            )
            i += 1

    return customers


# Generates products and prepares data for database insertion
def generate_products(colors, patterns, product_ids, subcategories_, prod_id=0):

    products_to_insert = []
    for subcategory_id, name, _ in subcategories_:
        for _ in range(5): 
            product_name = random.choice(adjectives) + " " + name
            color = random.choice(colors)
            pattern = random.choice(patterns)
            price = random.randint(30, 100)
            products_to_insert.append(
                (
                    product_ids[prod_id],
                    product_name,
                    color,
                    pattern,
                    price,
                    subcategory_id,
                )
            )

            prod_id += 1
    return products_to_insert


# Assigns random fabrics to the products and prepares the list of fabric details for database insertion
def generate_fabric_details(
    fabrics, subcategories_, product_ids, nr_prod, nr_prod_fabr=1
):
    fabric_details = []
    p_double_fabr = 0.1
    for i in range(1, nr_prod+1):
        rand_p = random.uniform(0, 1)
        rand_fabr = random.randint(1, len(fabrics))
        if rand_p > p_double_fabr:
            fabric_details.append((nr_prod_fabr, rand_fabr, product_ids[i - 1]))
            nr_prod_fabr += 1
        else:
            fabric_details.append((nr_prod_fabr, rand_fabr, product_ids[i - 1]))
            nr_prod_fabr += 1
            rand_fabr2 = random.randint(1, len(fabrics))
            while rand_fabr2 == rand_fabr:
                rand_fabr2 = random.randint(1, len(fabrics))
            fabric_details.append((nr_prod_fabr, rand_fabr2, product_ids[i - 1]))
            nr_prod_fabr += 1
    return fabric_details


# Generates dates between the start and end date
def generate_date(start_date, end_date):
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date.strftime("%Y-%m-%d"))
        current_date += datetime.timedelta(days=1)
    return dates


# Generates sales and prepares data for database insertion
def generate_sales(stores_, products_to_insert, dates, start_index=1, sd_id=1):
    sales_to_insert = []
    sale_details_to_insert = []
    for date in dates:
        num_sales_for_day = random.randint(5, 15)
        for _ in range(num_sales_for_day):
            customer = random.choice(customers_to_insert)
            customer_id = customer[0]
            store_id = random.randint(1, len(stores_))
            num_products = random.randint(1, 2)
            products_purchased = random.sample(products_to_insert, num_products)
            sum_price = 0
            for product in products_purchased:
                quantity = random.randint(1, 3)
                price = product[4] * quantity
                sum_price += price
                size = customer[6]
                size_index = sizes.index(size)
                if random.random() < 0.2 and size_index < len(sizes) - 1:
                    size = sizes[size_index + 1]
                sale_details_to_insert.append(
                    (sd_id, start_index, product[0], quantity, price, size)
                )
                sd_id += 1
            sales_to_insert.append((start_index, customer_id, store_id, date))
            start_index += 1
    return sales_to_insert, sale_details_to_insert


# Inserts all the generated data in MySQL database
def db_inserts():
    connection = mysql.connector.connect(
        host="localhost",
        database="clothingstore",
        user="clothing-user",
        password="clothing-pass",
    )

    mycursor = connection.cursor()

    insertCountry = """INSERT INTO countries(countryID, country) VALUES(%s, %s)"""

    insertCounties = (
        """INSERT INTO counties(countyID, county, countryID) VALUES(%s, %s, %s)"""
    )

    insertLocations = (
        """INSERT INTO locations(locationID, location, countyID) VALUES(%s, %s, %s)"""
    )

    insertFabrics = """INSERT INTO fabrics(fabricID, fabric) VALUES(%s, %s)"""

    insertFabricdetails = """INSERT INTO fabricdetails(fabricdetailsID, fabricID, productID) VALUES(%s, %s, %s)"""

    insertCustomers = """INSERT INTO customers(customerID, customerName, email, phoneNumber, gender, locationID, size) VALUES(%s, %s, %s, %s, %s, %s, %s)"""

    insertCategories = (
        """INSERT INTO categories(categoryID, category, gender) VALUES(%s, %s, %s)"""
    )

    insertSubcategories = """INSERT INTO subcategories(subcategoryID, subCategory, categoryID) VALUES(%s, %s, %s)"""

    insertStores = """INSERT INTO stores(storeID, storeName, locationID, address) VALUES(%s, %s, %s, %s)"""

    insertProducts = """INSERT INTO products(productID, productName, color, pattern, price, subcategoryID) VALUES(%s, %s, %s, %s, %s, %s)"""

    insertSales = """INSERT INTO sales(saleID, customerID, storeID, saleDate) VALUES(%s, %s, %s, %s)"""

    insertSaleDetails = """INSERT INTO saledetails(saleDetailID, saleID, productID, quantity, price, size) VALUES(%s, %s, %s, %s, %s, %s)"""

    mycursor.execute(insertCountry, (1, countries[0]))
    connection.commit()

    mycursor.executemany(insertCounties, counties_to_insert)
    connection.commit()

    mycursor.executemany(insertLocations, customer_locations_to_insert)
    connection.commit()

    mycursor.executemany(insertFabrics, fabrics_to_insert)
    connection.commit()

    mycursor.executemany(insertCategories, categories_to_insert)
    connection.commit()

    mycursor.executemany(insertSubcategories, subcategories_to_insert)
    connection.commit()

    mycursor.executemany(insertStores, stores_to_insert)
    connection.commit()

    mycursor.executemany(insertCustomers, customers_to_insert)
    connection.commit()

    mycursor.executemany(insertProducts, products_to_insert)
    connection.commit()

    mycursor.executemany(insertFabricdetails, fabric_details_to_insert)
    connection.commit()

    mycursor.executemany(insertSales, sales_to_insert)
    connection.commit()

    mycursor.executemany(insertSaleDetails, sale_details_to_insert)
    connection.commit()


# Creates and fills an Excel file with generated data
def excel_fill():
    global product_ids
    global fabric_details_to_insert
    global products_to_insert
    global subcategories_to_insert
    global customer_ids
    global customers_to_insert
    global sales_to_insert
    global sale_details_to_insert
    global customer_locations_to_insert
    global fabrics_to_insert
    global stores_to_insert
    global countries
    global location_ids

    # Create a new Workbook object
    workbook = openpyxl.Workbook()

    # Create sheets for different tables
    tables = [
        "countries",
        "counties",
        "locations",
        "fabrics",
        "categories",
        "subcategories",
        "stores",
        "customers",
        "products",
        "fabricdetails",
        "sales",
        "saledetails",
    ]

    countries = [(1, "Romania")]

    new_prod_cat_ids = [2, 4, 6]

    new_subcategories = [
        ["maternity top", "maternity sweatshirt"],
        ["maternity pants"],
        ["heels"],
    ]

    subcat_id = 0
    new_subcategories_to_insert = []
    index = len(subcategories_to_insert) + 1

    # Create new product categories
    for id in new_prod_cat_ids:
        for sc in new_subcategories[subcat_id]:
            new_subcategories_to_insert.append((index, sc, id))
            index += 1
        subcat_id += 1

    start_date = datetime.date(2021, 1, 1)
    end_date = datetime.date(2023, 12, 31)

    # Create new dates
    dates = generate_date(start_date, end_date)

    new_products_nr = len(new_subcategories_to_insert) * 5
    new_prod_ids = generate_id(new_products_nr, 5, product_ids)
    new_prod_id = len(products_to_insert)

    # Concatenate old product ids with new ones
    product_ids = product_ids + new_prod_ids

    # Generate the new products and prepare data for database insertion
    new_products = generate_products(
        colors, patterns, product_ids, new_subcategories_to_insert, new_prod_id
    )

    # Add new products to old ones
    products_to_insert += new_products

    # Assign random fabrics to new products and add the data to old fabric details
    new_fabr_details = generate_fabric_details(
        fabrics,
        new_subcategories_to_insert,
        new_prod_ids,
        new_products_nr,
        len(fabric_details_to_insert) + 1,
    )
    fabric_details = fabric_details_to_insert + new_fabr_details

    # Create new customers and add to old ones
    nr_customers = len(customers_to_insert)
    new_customer_ids = generate_id(nr_customers + 500, 15, customer_ids)
    customer_ids += new_customer_ids
    new_customers = generate_customers(nr_customers + 500, customer_ids, location_ids, nr_customers)
    customers_to_insert += new_customers

    # Create new sales
    new_start_id = len(sales_to_insert) + 1
    new_sd_id = len(sale_details_to_insert) + 1
    new_sales_to_insert, new_sale_details_to_insert = generate_sales(
        stores_to_insert, products_to_insert, dates, new_start_id, new_sd_id
    )

    subcategories_to_insert += new_subcategories_to_insert
    # Data for each table
    data = [
        countries,
        counties_to_insert,
        customer_locations_to_insert,
        fabrics_to_insert,
        categories_to_insert,
        subcategories_to_insert,
        stores_to_insert,
        customers_to_insert,
        products_to_insert,
        fabric_details,
        new_sales_to_insert,
        new_sale_details_to_insert,
    ]

    # Headers for each table
    headers = [
        ["countryID", "country"],
        ["countyID", "county", "countryID"],
        ["locationID", "location", "countyID"],
        ["fabricID", "fabric"],
        ["categoryID", "category", "gender"],
        ["subcategoryID", "subCategory", "categoryID"],
        ["storeID", "storeName", "locationID", "address"],
        [
            "cnp",
            "customerName",
            "email",
            "phoneNumber",
            "gender",
            "locationID",
            "size",
        ],
        ["productID", "productName", "color", "pattern", "price", "subcategoryID"],
        ["fabricdetailsID", "fabricID", "productID"],
        [
            "saleID",
            "customerID",
            "storeID",
            "saleDate",
        ],
        ["saleDetailsID", "saleID", "productID", "quantity", "price", "size"],
    ]

    # Fill data into sheets
    for table, table_data, header_row in zip(tables, data, headers):
        sheet = workbook.create_sheet(title=table)
        sheet.append(header_row)  # Adding headers

        for row in table_data:
            sheet.append(row)

    # Auto-adjust column width
    for sheet in workbook:
        for col in sheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

    # Bold header row
    for sheet in workbook:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    # Remove the default first sheet
    workbook.remove(workbook["Sheet"])

    # Save the workbook
    workbook.save("clothing_store_data.xlsx")


def main():

    global counties_to_insert
    global fabrics_to_insert
    global customer_locations
    global categories_to_insert
    global subcategories_to_insert
    global  locations_to_insert
    global customer_locations_to_insert
    global stores_to_insert
    global customers_to_insert
    global products_to_insert
    global fabric_details_to_insert
    global sales_to_insert
    global sale_details_to_insert
    global customer_ids
    global product_ids
    global location_ids

    customer_locations = ["location" + str(i) for i in range(1, 1000)]

    # Prepares the list of counties and fabrics for database insertion
    for i in range(1, len(counties) + 1):
        counties_to_insert.append((i, counties[i - 1], 1))

    for i in range(1, len(fabrics) + 1):
        fabrics_to_insert.append((i, fabrics[i - 1]))

    nr_dresses, categories_to_insert = generate_categories(categories)

    subcategories_to_insert = generate_subcategories(
        subcategories, nr_dresses, categories_to_insert
    )

    nr_store_locations = sum([len(k) for k in locations])

    # Generates IDs for store locations
    store_location_ids = generate_id(nr_store_locations, 6)

    # Generates IDs for customer locations
    customer_location_ids = generate_id(len(customer_locations), 6, store_location_ids)


    locations_to_insert, customer_locations_to_insert = generate_locations(
        counties, locations, customer_locations, store_location_ids, customer_location_ids
    )

    stores_to_insert = generate_stores(locations, store_location_ids)


    # Customer and location IDs generation
    customer_ids = generate_id(3000, 15)
    location_ids = store_location_ids + customer_location_ids

    customers_to_insert = generate_customers(3000, customer_ids, location_ids)

    nr_products = len(subcategories_to_insert) * 5
    product_ids = generate_id(nr_products, 5)


    products_to_insert = generate_products(
        colors, patterns, product_ids, subcategories_to_insert
    )


    fabric_details_to_insert = generate_fabric_details(
        fabrics, subcategories_to_insert, product_ids, nr_products
    )

    start_date = datetime.date(2018, 1, 1)
    end_date = datetime.date(2020, 12, 31)
    dates = generate_date(start_date, end_date)


    sales_to_insert, sale_details_to_insert = generate_sales(
        stores_to_insert, products_to_insert, dates
    )

    print("Filling up database")
    db_inserts()

    print("Filling up Excel file")
    excel_fill()



if __name__=="__main__": 
    main()